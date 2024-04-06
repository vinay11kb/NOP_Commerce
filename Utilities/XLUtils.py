import openpyxl

# file,sheetName
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]

    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rowno,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rowno,column=columnno).value=data
    workbook.save(file)

